# utils/xlsx.py
import json
import io
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import points_to_pixels, pixels_to_EMU # Added pixels_to_EMU
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker # Added OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
# Excel 대시보드 생성 함수
def create_dashboard_excel(llm_result, in_memory=False, username=None, dashboard_screenshot=None):
    """
    LLM 분석 결과를 기반으로 대시보드 Excel 파일을 생성합니다.
    
    Args:
        llm_result (dict or str): LLM 분석 결과
        in_memory (bool, optional): 메모리에 파일을 생성할지 여부. True면 파일을 디스크에 저장하지 않고 
                                    메모리에서 바이트 객체로 반환합니다.
        username (str, optional): 담당자 이름. 기본값은 None이며, 이 경우 ""으로 설정됩니다.
        dashboard_screenshot (BytesIO, optional): 대시보드 스크린샷 메모리 데이터
    
    Returns:
        Path 또는 BytesIO: in_memory가 False이면 생성된 Excel 파일 경로, True이면 메모리 상의 파일 객체
    """
    # [SB] 기본 설정 및 경로 생성
    logs_dir = Path("logs")
    
    # [SB] 디스크에 저장할 때만 logs 폴더 생성 (xlsx_test.py용)
    if not in_memory:
        logs_dir.mkdir(exist_ok=True)
    
    # [SB] 날짜 형식을 yyyymmdd로 변경
    formatted_date = datetime.now().strftime("%Y%m%d")
    
    # [SB] 파일명에 사용자 이름 포함
    user_str = ""
    if username:
        # [SB] 공백 제거 및 특수문자 제거
        clean_username = ''.join(c for c in username if c.isalnum() or c.isspace()).strip()
        user_str = f"_{clean_username}"
    
    excel_filename = f"당직체크리스트v5_{formatted_date}{user_str}.xlsx"
    excel_path = logs_dir / excel_filename
    screenshot_dir = Path("screenshot")
    
    # [SB] 결과 데이터 변환
    llm_data = json.loads(llm_result) if isinstance(llm_result, str) else llm_result
    
    # [SB] 담당자 이름 설정
    current_user = username if username else ""
    
    # [SB] 체크리스트 데이터 정의 (데이터는 이전과 동일하므로 생략)
    data = [
        # 1. 로그인
        {
            "카테고리": "backOffice",
            "메뉴": "로그인",
            "체크사항": "https://hydra2.uxcloud.net 접속\nID: bomanager",
            "페이지스크린샷": "1.jpg",
            "결과": "정상" if llm_data.get("로그인상태") == "예" else "오류"
        },
        # 2. 예치금
        {
            "카테고리": "backOffice",
            "메뉴": "예치금",
            "체크사항": "Whois 200.00 USD 이상\nGabia 200,000 KRW 이상\n예치금이 남았는 지 확인",
            "페이지스크린샷": "2.jpg",
            "결과": f"{llm_data.get('whois_usd')}\n{llm_data.get('gabia_krw')}"
        },
        # 3. 스케줄러
        {
            "카테고리": "backOffice",
            "메뉴": "스케줄러",
            "체크사항": "스케쥴러 상태가 적용중 인지 확인",
            "페이지스크린샷": "3.jpg",
            "결과": "적용됨" if llm_data.get("스케줄러상태") == "예" else "미적용 상태"
        },
        # 4. 고객문의 - 1:1 문의
        {
            "카테고리": "backOffice",
            "메뉴": "고객문의",
            "체크사항": "1:1 문의 답변 준비중이 있는지 확인",
            "페이지스크린샷": "4.jpg",
            "결과": "0개 확인" if llm_data.get("1:1문의") == "예" else "1:1 답변준비중 있음"
        },
        # 5. 고객문의 - 이메일 문의
        {
            "카테고리": "backOffice",
            "메뉴": "고객문의",
            "체크사항": "이메일 문의 답변 준비중이 있는지 확인",
            "페이지스크린샷": "5.jpg",
            "결과": "0개 확인" if llm_data.get("이메일문의") == "예" else "이메일 답변준비중 있음"
        },
        # 6. 고객문의 - 에러리포트
        {
            "카테고리": "backOffice",
            "메뉴": "고객문의",
            "체크사항": "에러리포트 신규 등록 된 이슈가 있는지 확인",
            "페이지스크린샷": "6.jpg",
            "결과": "0개 확인" if llm_data.get("에러리포트") == "예" else "신규 에러 있음"
        },
        # 7. Region
        {
            "카테고리": "backOffice",
            "메뉴": "Region",
            "체크사항": "Region 상태가 2개 활성화 인지 확인",
            "페이지스크린샷": "7.jpg",
            "결과": "2개 확인" if llm_data.get("Region활성") == "예" else "활성화 개수 확인 필요"
        },
        # 8. 시스템 - 장비 미보고
        {
            "카테고리": "backOffice",
            "메뉴": "시스템",
            "체크사항": "장비의 미보고가 있는지 확인",
            "페이지스크린샷": "8.jpg",
            "결과": "미보고 0개 확인" if llm_data.get("장비미보고") == "예" else "미보고 장비 있음"
        },
        # 9. 시스템 - DBSync
        {
            "카테고리": "backOffice",
            "메뉴": "시스템",
            "체크사항": "DBSync 일시중지 및 오류가 있는지 확인",
            "페이지스크린샷": "9.jpg",
            "결과": "0개 확인" if (llm_data.get("DB_Sync", {}).get("일시중지") == "예" and 
                         llm_data.get("DB_Sync", {}).get("오류") == "예") else "일시중지 또는 오류 있음"
        },
        # 10. 시스템 - FrontEnd
        {
            "카테고리": "backOffice",
            "메뉴": "시스템",
            "체크사항": "1. FrontEnd 서버 상태가 정상인지 확인\n※ 정상이 아닐경우 열기를 통해서 사이트 이동이 되는지 확인\n\n2. FrontEnd 도메인 검색이 정상적으로 가능한지 확인",
            "페이지스크린샷": "10.jpg",
            "결과": ("정상" if (llm_data.get("FrontEnd", {}).get("상태") == "예" and 
                        llm_data.get("FrontEnd", {}).get("도메인 검색") == "예") else "비정상") + \
                    ("" if "link" not in llm_data.get("FrontEnd", {}) else 
                    f"\n더보기 링크 : {llm_data.get('FrontEnd', {}).get('link')}")
        },
        # 11. 운영중인 서비스
        {
            "카테고리": "backOffice",
            "메뉴": "운영중인 서비스",
            "체크사항": "Parking, URL, FURL 모두 정상인지 확인\n※ 정상이 아닐경우 해당 상태를 클릭하여 페이지 이동이 정상적으로 되는지 확인",
            "페이지스크린샷": "11.jpg",
            "결과": ("정상" if (llm_data.get("운영중인서비스", {}).get("parking") == "예" and 
                            llm_data.get("운영중인서비스", {}).get("url") == "예" and
                            llm_data.get("운영중인서비스", {}).get("furl") == "예") else "비정상") + \
                    ("" if "parking_link" not in llm_data.get("운영중인서비스", {}) else 
                    f"\nParking 링크 : {llm_data.get('운영중인서비스', {}).get('parking_link')}") + \
                    ("" if "url_link" not in llm_data.get("운영중인서비스", {}) else 
                    f"\nURL 링크 : {llm_data.get('운영중인서비스', {}).get('url_link')}") + \
                    ("" if "furl_link" not in llm_data.get("운영중인서비스", {}) else 
                    f"\nFURL 링크 : {llm_data.get('운영중인서비스', {}).get('furl_link')}")
        }
    ]
    
    # [SB] 워크북 초기화 및 시트 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "대시보드 체크리스트"
    img_ws = wb.create_sheet(title="원본 스크린샷")
    
    # [SB] 스타일 정의
    styles = {
        'title': Font(size=14, bold=True),
        'header': Font(size=14, bold=True, color='FFFFFF'), # Header font color white
        'header_fill': PatternFill(start_color='003366', end_color='003366', fill_type='solid'),
        'good_result': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
        'bad_result': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
        'center_align': Alignment(horizontal='center', vertical='center'),
        'wrap_text': Alignment(wrapText=True, vertical='center', horizontal='center'),
        'thin_border': Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        ),
    }
    
    # [SB] 열 너비 설정
    column_widths = {'A': 12, 'B': 22, 'C': 50, 'D': 100, 'E': 20} # D 열 너비 100으로 유지
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # [SB] 타이틀 및 정보 행 설정
    ws.merge_cells('A1:E1')
    ws['A1'] = "당직 체크 리스트"
    ws['A1'].font = styles['title']
    ws['A1'].alignment = styles['center_align']
    
    ws.merge_cells('A2:C2')
    ws['A2'] = f"일시: {datetime.now().strftime('%Y-%m-%d')}"
    ws.merge_cells('D2:E2')
    ws['D2'] = f"담당자: {current_user}"
    
    # [SB] 헤더 행 설정
    headers = ['카테고리', '메뉴', '체크사항', '페이지스크린샷', '결과']
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = ws[f'{col_letter}3']
        cell.value = header
        cell.font = styles['header']
        cell.alignment = styles['center_align']
        cell.fill = styles['header_fill']
    
    # [SB] 카테고리별 행 인덱스 수집
    category_rows = {}
    for row_idx, row_data in enumerate(data, 4):
        category = row_data['카테고리']
        menu = row_data['메뉴']
        
        if category not in category_rows:
            category_rows[category] = []
        category_rows[category].append(row_idx)
        
        menu_key = f"{category}_{menu}"
        if menu_key not in category_rows:
            category_rows[menu_key] = []
        category_rows[menu_key].append(row_idx)
        
        ws[f'A{row_idx}'] = row_data['카테고리']
        ws[f'B{row_idx}'] = row_data['메뉴']
        ws[f'C{row_idx}'] = row_data['체크사항']
        
        result_cell = ws[f'E{row_idx}']
        result_cell.value = row_data['결과']
        
        menu = row_data['메뉴']
        result_text = row_data['결과']

        if menu == "로그인":
            if "정상" in result_text: result_cell.fill = styles['good_result']
            else: result_cell.fill = styles['bad_result']
        elif menu == "예치금":
            whois_ok, gabia_ok = False, False
            lines = result_text.split('\n')
            for line in lines:
                if 'USD' in line:
                    try:
                        usd_value = float(line.replace('USD', '').replace(',', '').strip())
                        whois_ok = usd_value >= 200.00
                    except ValueError: whois_ok = False
                elif 'KRW' in line:
                    try:
                        krw_value = float(line.replace('KRW', '').replace(',', '').strip())
                        gabia_ok = krw_value >= 200000
                    except ValueError: gabia_ok = False
            if whois_ok and gabia_ok: result_cell.fill = styles['good_result']
            else: result_cell.fill = styles['bad_result']
        elif menu == "스케줄러":
            if "적용됨" in result_text: result_cell.fill = styles['good_result']
            else: result_cell.fill = styles['bad_result']
        elif menu == "고객문의":
            if "0개 확인" in result_text: result_cell.fill = styles['good_result']
            else: result_cell.fill = styles['bad_result']
        elif menu == "Region":
            if "2개 확인" in result_text: result_cell.fill = styles['good_result']
            else: result_cell.fill = styles['bad_result']
        elif menu == "시스템":
            if "FrontEnd" in row_data['체크사항']:
                if (result_text.split('\n')[0] if result_text else "") == "정상":
                    result_cell.fill = styles['good_result']
                else: result_cell.fill = styles['bad_result']
            else:
                if "미보고 0개 확인" in result_text or "0개 확인" in result_text or "정상" in result_text:
                    result_cell.fill = styles['good_result']
                else: result_cell.fill = styles['bad_result']
        elif menu == "운영중인 서비스":
            if "정상" in result_text and "비정상" not in result_text:
                result_cell.fill = styles['good_result']
            else: result_cell.fill = styles['bad_result']
        
        ws.row_dimensions[row_idx].height = 120 # 행 높이 고정
    
    def merge_consecutive_rows(rows, column):
        if not rows: return
        groups, current_group = [], [rows[0]]
        for i in range(1, len(rows)):
            if rows[i] == rows[i-1] + 1: current_group.append(rows[i])
            else: groups.append(current_group); current_group = [rows[i]]
        groups.append(current_group)
        for group in groups:
            if len(group) > 1:
                ws.merge_cells(f'{column}{group[0]}:{column}{group[-1]}')
                ws[f'{column}{group[0]}'].alignment = styles['wrap_text']
    
    for category, rows in category_rows.items():
        if '_' in category: merge_consecutive_rows(rows, 'B')
        else: merge_consecutive_rows(rows, 'A')
    
    for row in range(1, len(data) + 4):
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.border = styles['thin_border']
            if col != 4: cell.alignment = styles['wrap_text']
            elif col == 4 and row > 3 : # D열 데이터 행에만 중앙 정렬 (이미지용)
                 cell.alignment = styles['center_align']


    # --- D셀에 이미지가 오버레이되어 D셀의 왼쪽, 위 테두리선이 보이지 않는 문제(구글 스프레드시트와 엑셀의 결과가 다르게 나타남) 해결을 위해 코드가 길어짐 ---
    for row_idx, row_data in enumerate(data, 4):
        try:
            menu = row_data['메뉴']
            result_text = row_data['결과']
            has_problem = False
            
            # [SB] 문제 여부 판단 로직 (기존과 동일)
            if menu == "로그인": 
                has_problem = "정상" not in result_text
            elif menu == "예치금":
                whois_ok, gabia_ok = False, False
                lines = result_text.split('\n')
                for line in lines:
                    if 'USD' in line:
                        try: 
                            usd_value = float(line.replace('USD', '').replace(',', '').strip())
                            whois_ok = usd_value >= 200.00
                        except ValueError: 
                            whois_ok = False
                    elif 'KRW' in line:
                        try: 
                            krw_value = float(line.replace('KRW', '').replace(',', '').strip())
                            gabia_ok = krw_value >= 200000
                        except ValueError: 
                            gabia_ok = False
                has_problem = not (whois_ok and gabia_ok)
            elif menu == "스케줄러": 
                has_problem = "적용됨" not in result_text
            elif menu == "고객문의": 
                has_problem = "0개 확인" not in result_text
            elif menu == "Region": 
                has_problem = "2개 확인" not in result_text
            elif menu == "시스템":
                if "FrontEnd" in row_data['체크사항']: 
                    has_problem = (result_text.split('\n')[0] if result_text else "") != "정상"
                else: 
                    has_problem = not ("미보고 0개 확인" in result_text or "0개 확인" in result_text or "정상" in result_text)
            elif menu == "운영중인 서비스": 
                has_problem = not ("정상" in result_text and "비정상" not in result_text)
            
            cell_D = ws[f'D{row_idx}']  # [SB] 대상 셀

            if has_problem:
                # [SB] 문제가 있는 경우, 텍스트 메시지 표시
                cell_D.value = "아래 원본 스크린샷 시트를 확인해주세요"
                cell_D.font = Font(color='FF0000', bold=True)
            else:
                # [SB] 문제가 없는 경우, 이미지 추가
                img_file = row_data['페이지스크린샷']
                img_path = screenshot_dir / img_file

                try:
                    img_object = Image(str(img_path))
                    
                    # [SB] 이미지별 개별 패딩 설정
                    if img_file == "1.jpg":  # [SB] 로그인 이미지만 특별 처리
                        padding_left = 20     # [SB] 왼쪽 패딩 줄임
                        padding_top = 10
                        padding_right = 30  # [SB] 오른쪽 패딩 늘림
                        padding_bottom = 10
                    else:  # [SB] 나머지 이미지들은 기존대로
                        padding_left = 10
                        padding_top = 10
                        padding_right = 10
                        padding_bottom = 10

                    # [SB] 셀 객체 및 크기 계산
                    target_cell_obj = ws.cell(row=row_idx, column=4)  # D열은 4번째 열
                    
                    # [SB] 셀 높이 및 너비 계산 (기존과 동일)
                    row_dim_obj = ws.row_dimensions[target_cell_obj.row]
                    cell_height_points = row_dim_obj.height if row_dim_obj.height is not None else 120.0 
                    cell_height_pixels = points_to_pixels(cell_height_points)

                    col_dim_obj = ws.column_dimensions[target_cell_obj.column_letter]
                    cell_width_char_units = col_dim_obj.width if col_dim_obj.width is not None else 100.0 
                    estimated_cell_width_pixels = cell_width_char_units * 7.0 
                    
                    # [SB] 개별 패딩을 적용한 실제 이미지 컨테이너 크기
                    img_container_width_px = estimated_cell_width_pixels - (padding_left + padding_right)
                    img_container_height_px = cell_height_pixels - (padding_top + padding_bottom)

                    if img_container_width_px <= 0 or img_container_height_px <= 0:
                        cell_D.value = "셀이 너무 작습니다"
                        cell_D.font = Font(color='FFA500')  # 주황색
                        continue

                    # [SB] 원본 이미지 크기 및 스케일링 계산
                    original_img_width_px = img_object.width
                    original_img_height_px = img_object.height
                    scaled_img_width_px = original_img_width_px
                    scaled_img_height_px = original_img_height_px

                    if original_img_width_px > 0 and original_img_height_px > 0:
                        # [SB] 비율 유지를 위한 스케일 팩터 계산
                        ratio_w = img_container_width_px / original_img_width_px
                        ratio_h = img_container_height_px / original_img_height_px
                        scale_factor = min(ratio_w, ratio_h)
                        
                        scaled_img_width_px = int(original_img_width_px * scale_factor)
                        scaled_img_height_px = int(original_img_height_px * scale_factor)
                    else: 
                        # [SB] 유효하지 않은 이미지 크기일 경우 기본값
                        scaled_img_width_px = 50 
                        scaled_img_height_px = 50
                    
                    # [SB] 1.jpg는 왼쪽 정렬, 나머지는 중앙 정렬
                    if img_file == "1.jpg":
                        # [SB] 1.jpg는 왼쪽 정렬 (중앙 정렬 계산 제거)
                        x_offset_px = padding_left  # [SB] 왼쪽 정렬
                        y_offset_px = padding_top + (img_container_height_px - scaled_img_height_px) / 2  # [SB] 세로는 중앙
                    else:
                        # [SB] 나머지 이미지들은 중앙 정렬
                        x_offset_px = padding_left + (img_container_width_px - scaled_img_width_px) / 2
                        y_offset_px = padding_top + (img_container_height_px - scaled_img_height_px) / 2

                    print(f"[SB] {img_file}: x_offset={x_offset_px:.1f}")

                    # [SB] 디버깅용 로그
                    print(f"[SB] {img_file}: left={padding_left}, x_offset={x_offset_px:.1f}")

                    # [SB] OneCellAnchor 생성 
                    anchor = OneCellAnchor()
                    
                    # [SB] _from 속성 설정 (AnchorMarker 사용)
                    anchor._from = AnchorMarker(col=target_cell_obj.column - 1, 
                                                colOff=pixels_to_EMU(x_offset_px),
                                                row=target_cell_obj.row - 1, 
                                                rowOff=pixels_to_EMU(y_offset_px))
                    
                    # [SB] 핵심 수정: ext를 XDRPositiveSize2D 객체로 설정
                    anchor.ext = XDRPositiveSize2D(cx=pixels_to_EMU(scaled_img_width_px), 
                                                cy=pixels_to_EMU(scaled_img_height_px))
                    
                    # [SB] 셀의 기존 내용을 지우고 이미지 추가
                    cell_D.value = None  # [SB] 이미지와 텍스트가 겹치지 않도록 셀 값 비우기
                    ws.add_image(img_object, anchor)

                except FileNotFoundError:
                    cell_D.value = f"이미지 파일 없음: {img_file}"
                    cell_D.font = Font(color='FF8C00') 
                except Exception as img_error:
                    # [SB] 상세 오류 메시지를 콘솔에 출력
                    print(f"[SB] Error processing image '{img_file}'. Path: '{img_path.resolve()}'. Error: {type(img_error).__name__} - {str(img_error)}")
                    cell_D.value = f"이미지 오류: {img_file}"
                    cell_D.font = Font(color='FF0000')              
        except Exception as e:
            ws[f'D{row_idx}'].value = f"행 처리 오류: {str(e)}"
            ws[f'D{row_idx}'].font = Font(color='800080') 
    
    if dashboard_screenshot:
        try:
            dashboard_screenshot.seek(0)
            dashboard_img = Image(dashboard_screenshot)
            img_ws.add_image(dashboard_img, 'A1')
            img_ws.row_dimensions[1].height = min(600, dashboard_img.height * 0.75)
            img_ws.column_dimensions['A'].width = min(150, dashboard_img.width * 0.1)
            print(f"[SB] 대시보드 스크린샷이 추가되었습니다")
        except Exception as e:
            print(f"[SB] 대시보드 스크린샷 추가 실패: {e}")
            img_ws['A1'] = f"스크린샷 오류: {e}"
    else:
        img_ws['A1'] = "대시보드 스크린샷이 제공되지 않았습니다."
    
    if in_memory:
        memory_file = io.BytesIO()
        wb.save(memory_file)
        memory_file.seek(0)
        memory_file.name = excel_filename
        print(f"[SB] Excel 파일이 메모리에 생성되었습니다: {excel_filename}")
        return memory_file
    else:
        wb.save(excel_path)
        print(f"[SB] Excel 파일이 디스크에 저장되었습니다: {excel_path}")
        return excel_path

